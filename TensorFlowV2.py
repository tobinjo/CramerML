from openpyxl import load_workbook
wb2 = load_workbook('CramerPicks.xlsx')
wb2.active = 8
ws = wb2.active
print ws.title
biglist = []
jimschoices = []
rowval = 2

for row in ws.iter_rows('D2:R3268'):
    #print rowval
    rowlist = []
    good = 1
    zeros = [0, 0, 0, 0, 0]
    choice = ws.cell(row = rowval, column = 3)
    #print choice.value
    zeros[choice.value-1] = 1
    for cell in row:
        if cell.value == '#N/A':
            good = 0
        rowlist.append(cell.value)
    if good:
        biglist.append(rowlist)
        jimschoices.append(zeros)
    rowval += 1

import numpy as np
import tensorflow as tf
Xtr = np.asanyarray(biglist[1:2499])
Ytr = np.asanyarray(jimschoices[1:2499])
Xte = np.asanyarray(biglist[2500:])
Yte = np.asanyarray(jimschoices[2500:])

x = tf.placeholder(tf.float32, [None, 15])
W = tf.Variable(tf.zeros([15, 5]))
b = tf.Variable(tf.zeros([5]))
y = tf.nn.softmax(tf.matmul(x, W) + b)

y_ = tf.placeholder(tf.float32, [None, 5])
cross_entropy = tf.reduce_mean(-tf.reduce_sum(y_ * tf.log(y), reduction_indices=[1]))
train_step = tf.train.GradientDescentOptimizer(0.5).minimize(cross_entropy)
init = tf.initialize_all_variables()
sess = tf.Session()
sess.run(init)

for i in range(200):
    batchxs = np.asanyarray(biglist[(i-1)*10:(i*10)-1])
    batchys = np.asanyarray(jimschoices[(i-1)*10:(i*10)-1])
    sess.run(train_step, feed_dict={x: batchxs, y_: batchys})

correct_prediction = tf.equal(tf.argmax(y,1), tf.argmax(y_,1))
accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))
print(sess.run(accuracy, feed_dict={x: Xte, y_: Yte}))
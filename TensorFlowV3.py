from openpyxl import load_workbook
wb2 = load_workbook('CramerPicks.xlsx')
wb2.active = 8
ws = wb2.active
biglist = []
jimschoices = []
rowval = 2

for row in ws.iter_rows('D2:R3268'):
    rowlist = []
    good = 1
    zeros = [0, 0, 0, 0, 0]
    choice = ws.cell(row = rowval, column = 3)
    zeros[choice.value-1] = 1
    for cell in row:
        if cell.value == '#N/A':
            good = 0
        rowlist.append(cell.value)
    if good:
        biglist.append(rowlist)
        jimschoices.append(zeros)
    rowval += 1

import numpy as np
import tensorflow as tf
Xtr = np.asanyarray(biglist[0:2499])
Ytr = np.asanyarray(jimschoices[0:2499])
Xte = np.asanyarray(biglist[2500:])
Yte = np.asanyarray(jimschoices[2500:])

# Parameters
learning_rate = 0.01
training_epochs = 30
batch_size = 100
display_step = 1

# Network Parameters
n_hidden_1 = 32 # 1st layer number of features
n_hidden_2 = 16 # 2nd layer number of features
n_input = 15
n_classes = 5

# tf Graph input
x = tf.placeholder("float", [None, n_input])
y = tf.placeholder("float", [None, n_classes])

# Create model
def multilayer_perceptron(x, weights, biases):
    # Hidden layer with RELU activation
    layer_1 = tf.add(tf.matmul(x, weights['h1']), biases['b1'])
    layer_1 = tf.nn.relu(layer_1)
    # Hidden layer with RELU activation
    layer_2 = tf.add(tf.matmul(layer_1, weights['h2']), biases['b2'])
    layer_2 = tf.nn.relu(layer_2)
    # Output layer with linear activation
    out_layer = tf.matmul(layer_2, weights['out']) + biases['out']
    return out_layer

# Store layers weight & bias
weights = {
    'h1': tf.Variable(tf.random_normal([n_input, n_hidden_1])),
    'h2': tf.Variable(tf.random_normal([n_hidden_1, n_hidden_2])),
    'out': tf.Variable(tf.random_normal([n_hidden_2, n_classes]))
}
biases = {
    'b1': tf.Variable(tf.random_normal([n_hidden_1])),
    'b2': tf.Variable(tf.random_normal([n_hidden_2])),
    'out': tf.Variable(tf.random_normal([n_classes]))
}

# Construct model
pred = multilayer_perceptron(x, weights, biases)

# Define loss and optimizer
cost = tf.reduce_mean(tf.nn.softmax_cross_entropy_with_logits(pred, y))
optimizer = tf.train.AdamOptimizer(learning_rate=learning_rate).minimize(cost)

init = tf.initialize_all_variables()

with tf.Session() as sess:
    sess.run(init)
    
    # Training cycle
    for epoch in range(training_epochs):
        avg_cost = 0.
        total_batch = int(2500/batch_size)
        # Loop over all batches
        for i in range(total_batch):
            batch_xs = Xtr[i*100:(i*100)+100]
            batch_ys = Ytr[i*100:(i*100)+100]
            # Fit training using batch data
            _, c = sess.run([optimizer, cost], feed_dict={x: batch_xs,
                            y: batch_ys})
                            # Compute average loss
            avg_cost += c / total_batch
        # Display logs per epoch step
        if (epoch+1) % display_step == 0:
            print "Epoch:", '%04d' % (epoch+1), "cost=", "{:.9f}".format(avg_cost)

    print "Optimization Finished!"
    
    # Test model
    correct_prediction = tf.equal(tf.argmax(pred, 1), tf.argmax(y, 1))
    # Calculate accuracy for 3000 examples
    accuracy = tf.reduce_mean(tf.cast(correct_prediction, tf.float32))
    print "Accuracy:", accuracy.eval({x: Xte, y: Yte})